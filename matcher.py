"""
matcher.py - The RAG matching engine.

Contains the CodeMatcher class that:
1. Retrieves top-K candidates from ChromaDB
2. Sends them to the LLM with the matching prompt
3. Parses the structured JSON response
"""

import json
import time
import asyncio
import pandas as pd
import config
from prompts import MATCH_PROMPT, NORMALIZATION_PROMPT, format_candidates

from pinecone import Pinecone

class CodeMatcher:
    """RAG-based code matching engine."""

    def __init__(self):
        """Load vector store and LLM based on configuration."""
        
        # Initialize Pinecone
        self.pc = Pinecone(api_key=config.PINECONE_API_KEY)
        self.index = self.pc.Index(config.PINECONE_INDEX_NAME)

        # Initialize LLM (AWS Bedrock)
        import boto3
        from botocore.config import Config
        
        # Configure retry strategy
        boto_config = Config(
            retries={
                'max_attempts': 5,
                'mode': 'adaptive'
            }
        )
        
        self.bedrock_client = boto3.client(
            service_name="bedrock-runtime",
            region_name=config.AWS_REGION,
            config=boto_config
        )
        self.llm_model_id = config.LLM_MODEL
        self.use_boto = True
        
        # Initialize cache for performance optimization
        self.cache = {}
        self.cache_hits = 0
        self.cache_misses = 0

    def _extract_metadata(self, doc) -> dict:
        """Extract system-specific metadata from a document."""
        metadata = doc.metadata
        system = metadata.get("system", "")
        
        result = {
            "matched_code": metadata.get("code", ""),
            "code_system": system,
            "matched_description": metadata.get("description", ""),
        }
        
        # For SBS, add additional fields
        if system == "SBS":
            result["sbs_code_numeric"] = metadata.get("code_numeric", "")
            result["sbs_code_hyphenated"] = metadata.get("code_hyphenated", "")
            result["short_description"] = metadata.get("description", "")
        
        # For GTIN, add additional fields
        elif system == "GTIN":
            result["gtin_code"] = metadata.get("code", "")
            result["gtin_ingredients"] = metadata.get("ingredients", "")
            result["gtin_strength"] = metadata.get("strength", "")
        
        # For GMDN, add additional fields
        elif system == "GMDN":
            result["gmdn_code"] = metadata.get("code", "")
            result["gmdn_name"] = metadata.get("term_name", "")
            result["gmdn_definition"] = metadata.get("term_definition", "")
        
        return result
    
    def clear_cache(self):
        """Clear the match cache."""
        self.cache = {}
        self.cache_hits = 0
        self.cache_misses = 0
    
    def get_cache_stats(self) -> dict:
        """Get cache statistics."""
        total = self.cache_hits + self.cache_misses
        hit_rate = (self.cache_hits / total * 100) if total > 0 else 0
        return {
            "cache_size": len(self.cache),
            "cache_hits": self.cache_hits,
            "cache_misses": self.cache_misses,
            "hit_rate": f"{hit_rate:.1f}%"
        }

    def normalize_query(self, user_input: str) -> dict:
        """
        Normalize user input using LLM to get standardized query and domain guess.
        """
        prompt = NORMALIZATION_PROMPT.format(service_description=user_input)
        
        response_text = None
        # Call LLM (reuse existing LLM logic or simplify)
        if self.use_boto:
             payload = {
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 1024,
                "messages": [{"role": "user", "content": [{"type": "text", "text": prompt}]}],
                "temperature": getattr(config, "NORMALIZATION_LLM_TEMPERATURE", 0.0)
            }
             try:
                model_id = getattr(config, "NORMALIZATION_LLM_MODEL", "anthropic.claude-3-haiku-20240307-v1:0")
                response = self.bedrock_client.invoke_model(
                    body=json.dumps(payload),
                    modelId=model_id,
                    accept="application/json",
                    contentType="application/json"
                )
                response_body = json.loads(response.get("body").read())
                response_text = response_body["content"][0]["text"]
             except Exception as e:
                 print(f"Normalization failed: {e}")
                 return {"normalized_query": user_input, "domain": None}
        else:
            try:
                response = self.llm.invoke(prompt)
                response_text = response.content if hasattr(response, 'content') else str(response)
            except Exception as e:
                print(f"Normalization failed: {e}")
                return {"normalized_query": user_input, "domain": None}

        # Parse response
        normalized_query = user_input
        domain_guess = None
        
        try:
            lines = response_text.split('\n')
            for i, line in enumerate(lines):
                if "Normalized Query:" in line:
                    normalized_query = lines[i+1].strip()
                elif "Coding Domain Guess:" in line:
                    domain_guess = lines[i+1].strip().lower()
                    # Clean up domain guess
                    if "sbs" in domain_guess: domain_guess = "sbs_v2"
                    elif "gmdn" in domain_guess: domain_guess = "gmdn_v2"
                    elif "gtin" in domain_guess: domain_guess = "gtin_v2"
        except Exception:
            pass
            
                
        return {
            "normalized_query": normalized_query,
            "domain": domain_guess,
            "original_response": response_text 
        }

    def retrieve_candidates(self, query: str, predicted_domain: str = None) -> list:
        """
        Retrieve candidates from Pinecone with namespace routing using Pinecone Inference.
        """
        # Generate embedding using Pinecone Inference API
        model = getattr(config, "PINECONE_INFERENCE_MODEL", "multilingual-e5-large") 
        try:
            embeddings_response = self.pc.inference.embed(
                model=model,
                inputs=[query],
                parameters={"input_type": "query"}
            )
            query_embedding = embeddings_response[0]['values']
        except Exception as e:
            print(f"Pinecone Inference Failed: {e}")
            return []
        
        results = []
        
        # Strategy: 
        # 1. Search Predicted Namespace (Top High)
        # 2. Search Other Namespaces (Top Low)

        namespaces = getattr(config, "PINECONE_NAMESPACES", [])
        if predicted_domain and predicted_domain in namespaces:
            primary_ns = predicted_domain
            secondary_ns = [ns for ns in namespaces if ns != primary_ns]
        else:
            primary_ns = None
            secondary_ns = namespaces # Search all equally if no prediction
            
        # Helper to process Pinecone match
        def process_matches(matches, system_override=None):
            docs = []
            for match in matches:
                meta = match['metadata']
                
                # Determine system
                system = meta.get('system')
                if not system:
                    if system_override:
                        system = system_override.replace("_v2", "").upper()
                    elif meta.get("sbs_code") or meta.get("sbs_code_hyphenated"):
                        system = "SBS"
                    elif meta.get("gtin_code"):
                        system = "GTIN"
                    elif meta.get("gmdn_code"):
                        system = "GMDN"
                
                meta['system'] = system
                
                # Normalize 'code' and 'description' based on system
                if system == "SBS":
                    meta['code'] = meta.get('sbs_code_hyphenated') or str(meta.get('sbs_code', '')).replace('.0', '')
                    meta['description'] = meta.get('long_description') or meta.get('short_description', '')
                    # Additional SBS fields
                    meta['code_numeric'] = str(meta.get('sbs_code', '')).replace('.0', '')
                    meta['code_hyphenated'] = meta.get('sbs_code_hyphenated', '')
                    
                elif system == "GTIN":
                    meta['code'] = meta.get('gtin_code', '')
                    meta['description'] = meta.get('display', '')
                    # Additional GTIN fields
                    meta['ingredients'] = meta.get('ingredients', '')
                    meta['strength'] = meta.get('strength', '')
                    
                elif system == "GMDN":
                    meta['code'] = meta.get('gmdn_code', '')
                    meta['description'] = meta.get('term_name', '')
                    # Additional GMDN fields
                    meta['term_name'] = meta.get('term_name', '')
                    meta['term_definition'] = meta.get('term_definition', '')

                # Fallback for display text if no description found
                text_content = meta.get('text') or meta.get('description', '') or str(meta)

                # Simple Document class to replace LangChain
                class Document:
                    def __init__(self, page_content, metadata=None):
                        self.page_content = page_content
                        self.metadata = metadata or {}

                doc = Document(page_content=text_content, metadata=meta)
                doc.metadata['score'] = match['score']
                docs.append(doc)
            return docs

        # Execute Searches
        import concurrent.futures
        
        search_tasks = []
        
        # Primary Search
        if primary_ns:
            k_primary = getattr(config, "ROUTING_PRIMARY_K", 15)
            try:
                res = self.index.query(
                    vector=query_embedding,
                    top_k=k_primary,
                    namespace=primary_ns,
                    include_metadata=True
                )
                results.extend(process_matches(res['matches'], primary_ns))
            except Exception as e:
                print(f"Error searching namespace {primary_ns}: {e}")

        # Secondary Searches (Safety Net)
        k_secondary = getattr(config, "ROUTING_SECONDARY_K", 3)
        for ns in secondary_ns:
             try:
                res = self.index.query(
                    vector=query_embedding,
                    top_k=k_secondary,
                    namespace=ns,
                    include_metadata=True
                )
                results.extend(process_matches(res['matches'], ns))
             except Exception as e:
                print(f"Error searching namespace {ns}: {e}")
                
        # Deduplicate results based on code+system
        seen = set()
        unique_results = []
        # Sort by score descending
        results.sort(key=lambda x: x.metadata.get('score', 0), reverse=True)
        
        for doc in results:
            key = f"{doc.metadata.get('system')}:{doc.metadata.get('code')}"
            if key not in seen:
                seen.add(key)
                unique_results.append(doc)
                
        return unique_results[:config.MAX_CANDIDATES_TO_LLM]

    def match_single(self, service_description: str) -> dict:
        """
        Match a single service description to a billing code.

        Pipeline:
        1. Validate input
        2. Check cache
        3. Expand query (abbreviations → formal SBS terminology)
        4. Two-pass retrieval (general + specialty-filtered)
        5. Format candidates with enriched metadata
        6. LLM evaluation with metadata-aware prompt
        7. Post-LLM validation
        8. Cache and return

        Args:
            service_description: The input service description to match.

        Returns:
            Dictionary with matched_code, code_system, matched_description,
            confidence, and reasoning.
        """
        # Handle empty or NaN service descriptions
        if not service_description or pd.isna(service_description):
            return {
                "matched_code": None,
                "code_system": None,
                "matched_description": None,
                "confidence": "NONE",
                "reasoning": "Empty or invalid service description",
            }

        service_description = str(service_description).strip()
        if not service_description or service_description.lower() == 'nan':
            return {
                "matched_code": None,
                "code_system": None,
                "matched_description": None,
                "confidence": "NONE",
                "reasoning": "Empty or invalid service description",
            }

        # Check cache
        cache_key = service_description.strip().lower()
        if cache_key in self.cache:
            self.cache_hits += 1
            return self.cache[cache_key].copy()

        self.cache_misses += 1

        # Step 1: Normalization & Routing (New Arch)
        normalization_result = self.normalize_query(service_description)
        mined_query = normalization_result["normalized_query"]
        predicted_domain = normalization_result["domain"]
        
        # Step 2: Namespace-aware Retrieval
        # Use refined query + original input for hybrid search signal? 
        # Prompt only asked to rewrite, but let's trust the rewrite for retrieval 
        # while keeping original for final LLM check.
        docs = self.retrieve_candidates(mined_query, predicted_domain)

        if not docs:
            result = {
                "matched_code": None,
                "code_system": None,
                "matched_description": None,
                "confidence": "NONE",
                "reasoning": "No candidates found in vector database",
            }
            self.cache[cache_key] = result
            return result

        top_doc = docs[0]

        # Step 3: Build prompt with ORIGINAL description (not expanded)
        # and enriched candidates (with full metadata)
        candidates_text = format_candidates(docs)
        prompt = MATCH_PROMPT.format(
            service_description=service_description,  # Original, not expanded
            candidates=candidates_text,
        )

        # Step 4: Call LLM with retry logic
        max_retries = 5
        base_delay = 5

        response_text = None
        
        # Prepare execution based on the selected LLM (Boto3 vs Local)
        if self.use_boto:
            # AWS Bedrock via Boto3
            payload = {
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 4096,
                "messages": [
                    {
                        "role": "user",
                        "content": [{"type": "text", "text": prompt}]
                    }
                ],
                "temperature": config.LLM_TEMPERATURE
            }
            body = json.dumps(payload)
            
            import botocore
            
            for attempt in range(max_retries):
                try:
                    response = self.bedrock_client.invoke_model(
                        body=body,
                        modelId=self.llm_model_id,
                        accept="application/json",
                        contentType="application/json"
                    )
                    
                    response_body = json.loads(response.get("body").read())
                    response_text = response_body["content"][0]["text"]
                    break
                    
                except botocore.exceptions.ClientError as e:
                    error_code = e.response['Error']['Code']
                    if error_code in ["ThrottlingException", "TooManyRequestsException"]:
                        if attempt < max_retries - 1:
                            delay = base_delay * (2 ** attempt)
                            print(f"Rate limited (AWS), retrying in {delay}s... (attempt {attempt + 1}/{max_retries})")
                            time.sleep(delay)
                            continue
                        else:
                            print(f"Max retries reached for throttling")
                            result = {
                                "matched_code": None,
                                "code_system": None,
                                "matched_description": None,
                                "confidence": "NONE",
                                "reasoning": "Rate limit exceeded (AWS), please try again later",
                            }
                            self.cache[cache_key] = result
                            return result
                    else:
                        print(f"AWS SDK Error: {e}")
                        result = self._extract_metadata(top_doc)
                        result["confidence"] = "LOW"
                        result["reasoning"] = f"AWS call failed: {str(e)[:100]}; using vector similarity"
                        self.cache[cache_key] = result
                        return result
                        
                except Exception as e:
                        print(f"Unexpected Error: {e}")
                        result = self._extract_metadata(top_doc)
                        result["confidence"] = "LOW"
                        result["reasoning"] = f"Call failed: {str(e)[:100]}; using vector similarity"
                        self.cache[cache_key] = result
                        return result

        else:
            # Local Ollama
            for attempt in range(max_retries):
                try:
                    response = self.llm.invoke(prompt)
                    if hasattr(response, 'content'):
                        response_text = response.content
                    else:
                        response_text = str(response)
                    break
                except Exception as e:
                    error_msg = str(e)
                    if "usage limit" in error_msg.lower() or "too many requests" in error_msg.lower():
                        if attempt < max_retries - 1:
                            delay = base_delay * (2 ** attempt)
                            print(f"Rate limited (Local), retrying in {delay}s... (attempt {attempt + 1}/{max_retries})")
                            time.sleep(delay)
                            continue
                    
                    print(f"Local LLM Error: {error_msg}")
                    result = self._extract_metadata(top_doc)
                    result["confidence"] = "LOW"
                    result["reasoning"] = f"LLM call failed: {error_msg[:100]}; using vector similarity"
                    self.cache[cache_key] = result
                    return result

        if response_text is None:
            result = self._extract_metadata(top_doc)
            result["confidence"] = "LOW"
            result["reasoning"] = "LLM call failed after retries; using vector similarity"
            self.cache[cache_key] = result
            return result

        # Step 5: Parse LLM response
        result = self._parse_llm_response(response_text, top_doc, docs)

        # Step 6: Post-LLM validation (safety net)
        if getattr(config, 'ENABLE_POST_LLM_VALIDATION', False):
            result = self._post_validate(result, service_description, docs)

        # Cache and return
        result["candidates"] = [
            {
                "code": doc.metadata.get("code"),
                "system": doc.metadata.get("system"),
                "description": doc.metadata.get("description"),
                "score": doc.metadata.get("score", "N/A") 
            } 
            for doc in docs
        ]
        self.cache[cache_key] = result
        return result

    def _parse_llm_response(self, response_text: str, fallback_doc, candidates: list = None) -> dict:
        """
        Parse the LLM's JSON response with fallback handling.

        Args:
            response_text: The raw text response from the LLM.
            fallback_doc: The top retrieval document to use as fallback.
            candidates: List of candidate documents retrieved from vector store.

        Returns:
            Parsed result dictionary enriched with system-specific fields.
        """
        result = None
        
        # DEBUG: Uncomment these lines if you need to troubleshoot LLM responses
        # print("\n" + "="*60)
        # print("DEBUG: Raw LLM Response:")
        # print("-"*60)
        # print(response_text[:500] if response_text else "EMPTY RESPONSE")
        # print("="*60 + "\n")
        
        try:
            result = json.loads(response_text)
        except json.JSONDecodeError as e:
            # print(f"DEBUG: Initial JSON parse failed: {e}")
            # Fallback: try stripping markdown code fences
            cleaned = response_text.strip()
            if cleaned.startswith("```json"):
                cleaned = cleaned[7:]
            elif cleaned.startswith("```"):
                cleaned = cleaned[3:]
            if cleaned.endswith("```"):
                cleaned = cleaned[:-3]
            cleaned = cleaned.strip()

            try:
                result = json.loads(cleaned)
                # print("DEBUG: Successfully parsed after cleaning markdown fences")
            except json.JSONDecodeError as e2:
                # print(f"DEBUG: Cleaned JSON parse also failed: {e2}")
                # print(f"DEBUG: Cleaned text (first 200 chars): {cleaned[:200]}")
                pass

        # If parsing failed, use fallback
        if result is None:
            # print("DEBUG: Using fallback (vector similarity)")
            result = self._extract_metadata(fallback_doc)
            result["confidence"] = "LOW"
            result["reasoning"] = "LLM response parsing failed; using vector similarity"
            return result
        
        # print(f"DEBUG: Successfully parsed JSON. Matched code: {result.get('matched_code')}")
        
        # Enrich the LLM result with system-specific fields by matching against candidates
        if candidates and result.get("matched_code"):
            matched_code = result.get("matched_code")
            code_system = result.get("code_system")
            
            # Find the matching document in candidates
            for doc in candidates:
                if (doc.metadata.get("code") == matched_code and 
                    doc.metadata.get("system") == code_system):
                    # Enrich with system-specific fields
                    if code_system == "SBS":
                        result["sbs_code_numeric"] = doc.metadata.get("code_numeric", "")
                        result["sbs_code_hyphenated"] = doc.metadata.get("code_hyphenated", "")
                        result["short_description"] = doc.metadata.get("description", "")
                    elif code_system == "GTIN":
                        result["gtin_code"] = doc.metadata.get("code", "")
                        result["gtin_ingredients"] = doc.metadata.get("ingredients", "")
                        result["gtin_strength"] = doc.metadata.get("strength", "")
                    elif code_system == "GMDN":
                        result["gmdn_code"] = doc.metadata.get("code", "")
                        result["gmdn_name"] = doc.metadata.get("term_name", "")
                        result["gmdn_definition"] = doc.metadata.get("term_definition", "")
                    break
        
        return result

    def _post_validate(self, result: dict, original_input: str, candidates: list) -> dict:
        """
        Post-LLM validation safety net. Catches remaining LLM errors by
        applying hard rules that override the LLM's decision.

        Rules:
        1. If input contains dental abbreviations but match is non-dental → NONE
        2. If matched code's Excludes field matches the input → NONE

        Args:
            result: The LLM's parsed result dict.
            original_input: The original (unexpanded) service description.
            candidates: The candidate documents used for matching.

        Returns:
            Possibly modified result dict.
        """
        if not result.get("matched_code"):
            return result  # Already null, nothing to validate

        input_lower = original_input.lower()
        matched_code = result.get("matched_code", "")
        code_system = result.get("code_system", "")

        # Rule 1: Dental abbreviation → non-dental code
        dental_signals = ["rct", "r.c.t", "root canal", "crown", "filling",
                          "extraction", "dental", "tooth", "pfm", "veneer",
                          "composite", "amalgam", "endodon", "pulp"]
        input_is_dental = any(s in input_lower for s in dental_signals)

        if input_is_dental and code_system == "SBS":
            # Find the matched document to check its chapter
            for doc in candidates:
                if doc.metadata.get("code") == matched_code:
                    chapter = doc.metadata.get("chapter_name", "").upper()
                    if chapter and "DENTAL" not in chapter and "ORAL" not in chapter:
                        result["matched_code"] = None
                        result["code_system"] = None
                        result["matched_description"] = None
                        result["confidence"] = "NONE"
                        result["reasoning"] = (
                            f"Post-validation rejected: dental input matched to "
                            f"non-dental chapter ({chapter})"
                        )
                        return result
                    break

        # Rule 2: Check Excludes field
        for doc in candidates:
            if doc.metadata.get("code") == matched_code:
                page_content = doc.page_content.lower()
                # Extract excludes text from enriched document
                if "excludes:" in page_content:
                    excludes_start = page_content.index("excludes:") + len("excludes:")
                    # Find the next section or end of string
                    excludes_text = page_content[excludes_start:]
                    next_section = excludes_text.find(". guideline:")
                    if next_section == -1:
                        next_section = len(excludes_text)
                    excludes_text = excludes_text[:next_section].strip()

                    # Check if input matches excludes content
                    input_words = set(input_lower.split())
                    excludes_words = set(excludes_text.split())
                    overlap = input_words & excludes_words
                    # If significant overlap with excludes, reject
                    meaningful_overlap = overlap - {"the", "a", "an", "of", "and", "or", "in", "to", "for"}
                    if len(meaningful_overlap) >= 3:
                        result["matched_code"] = None
                        result["code_system"] = None
                        result["matched_description"] = None
                        result["confidence"] = "NONE"
                        result["reasoning"] = (
                            f"Post-validation rejected: input matches Excludes field "
                            f"(overlap: {', '.join(list(meaningful_overlap)[:5])})"
                        )
                        return result
                break

        return result

    def match_batch(self, input_excel_path: str, output_excel_path: str) -> list:
        """
        Process an entire mapping sheet.

        Args:
            input_excel_path: Path to the input Excel file with Service Code
                              and Service Description columns.
            output_excel_path: Path where the output Excel will be saved.

        Returns:
            List of tuples (sheet_name, DataFrame) with results.
        """
        import os
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)

        xls = pd.ExcelFile(input_excel_path)
        all_results = []
        stats = {}

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)

            # Strip column names to handle trailing spaces
            df.columns = df.columns.str.strip()

            # Only keep Service Code and Service Description
            if "Service Code" not in df.columns or "Service Description" not in df.columns:
                print(f"Warning: Sheet '{sheet_name}' missing required columns, skipping")
                continue

            print(f"\nProcessing sheet: {sheet_name} ({len(df)} rows)")

            matched_codes = []
            code_systems = []
            matched_descriptions = []
            confidences = []
            reasonings = []
            
            # SBS-specific fields
            sbs_code_numeric_list = []
            sbs_short_desc_list = []
            sbs_code_hyphenated_list = []
            
            # GTIN-specific fields
            gtin_code_list = []
            gtin_ingredients_list = []
            gtin_strength_list = []
            

            gmdn_code_list = []
            gmdn_name_list = []
            gmdn_definition_list = []

            # Initialize stats for this sheet
            sheet_stats = {"total": len(df), "HIGH": 0, "MEDIUM": 0, "LOW": 0, "NONE": 0}

            # Process each row
            import time
            start_time = time.time()
            
            for idx, row in df.iterrows():
                desc = row.get("Service Description", "")
                desc_preview = str(desc)[:50] if desc else "(empty)"
                
                # Calculate progress
                progress = ((idx + 1) / len(df)) * 100
                elapsed = time.time() - start_time
                
                # Estimate time remaining
                if idx > 0:
                    avg_time_per_row = elapsed / (idx + 1)
                    remaining_rows = len(df) - (idx + 1)
                    eta_seconds = avg_time_per_row * remaining_rows
                    eta_minutes = eta_seconds / 60
                    
                    print(f"  [{sheet_name}] [{idx+1}/{len(df)}] ({progress:.1f}%) | ETA: {eta_minutes:.1f}m | {desc_preview}...")
                else:
                    print(f"  [{sheet_name}] [{idx+1}/{len(df)}] ({progress:.1f}%) | {desc_preview}...")

                result = self.match_single(desc)

                matched_codes.append(result.get("matched_code", ""))
                code_systems.append(result.get("code_system", ""))
                matched_descriptions.append(result.get("matched_description", ""))
                confidences.append(result.get("confidence", "NONE"))
                reasonings.append(result.get("reasoning", ""))
                
                # Collect SBS-specific fields
                sbs_code_numeric_list.append(result.get("sbs_code_numeric", ""))
                sbs_short_desc_list.append(result.get("short_description", ""))
                sbs_code_hyphenated_list.append(result.get("sbs_code_hyphenated", ""))
                
                # Collect GTIN-specific fields
                gtin_code_list.append(result.get("gtin_code", ""))
                gtin_ingredients_list.append(result.get("gtin_ingredients", ""))
                gtin_strength_list.append(result.get("gtin_strength", ""))
                
                # Collect GMDN-specific fields
                gmdn_code_list.append(result.get("gmdn_code", ""))
                gmdn_name_list.append(result.get("gmdn_name", ""))
                gmdn_definition_list.append(result.get("gmdn_definition", ""))

                # Update stats
                confidence = result.get("confidence", "NONE")
                if confidence in sheet_stats:
                    sheet_stats[confidence] += 1

                # Rate limiting to avoid API throttling
                delay = getattr(config, "MATCH_BATCH_DELAY", 5.0)
                time.sleep(delay)

            # Add new columns to DataFrame in the requested order
            # For SBS: SBS Code (numeric), Short Description, SBS Code Hyphenated
            df["SBS Code"] = sbs_code_numeric_list
            df["Short Description"] = sbs_short_desc_list
            df["SBS Code Hyphenated"] = sbs_code_hyphenated_list
            
            # For GTIN: GTIN Code, Ingredients, Strength
            df["GTIN Code"] = gtin_code_list
            df["GTIN Ingredients"] = gtin_ingredients_list
            df["GTIN Strength"] = gtin_strength_list
            
            # For GMDN: GMDN Code, Name, Definition
            df["GMDN Code"] = gmdn_code_list
            df["GMDN Name"] = gmdn_name_list
            df["GMDN Definition"] = gmdn_definition_list
            
            # Then add the common columns
            df["Matched Code"] = matched_codes
            df["Code System"] = code_systems
            df["Matched Description"] = matched_descriptions
            df["Confidence"] = confidences
            df["Reasoning"] = reasonings

            # Drop the empty NPHIES columns if they exist
            cols_to_drop = ["NPHIES Code", "Description", "Other Code Value"]
            for col in cols_to_drop:
                if col in df.columns:
                    df = df.drop(columns=[col])

            all_results.append((sheet_name, df))
            stats[sheet_name] = sheet_stats

        # Combine all results
        from openpyxl.styles import PatternFill

        # Combine all dataframes
        all_dfs = [df for _, df in all_results]
        if not all_dfs:
            print("No data to save.")
            return all_results
            
        combined_df = pd.concat(all_dfs, ignore_index=True)
        
        # --- Map columns to standardized output structure ---
        def get_service_type(row):
            sys = str(row.get("Code System", "")).strip()
            conf = str(row.get("Confidence", "")).upper()
            code = str(row.get("Matched Code", "")).strip()
            
            if not sys or conf == "NONE" or not code or sys.lower() == "nan":
                return "undefined"
            return sys

        def get_nphies_code(row):
            sys = str(row.get("Code System", "")).upper()
            if sys == "SBS": return row.get("SBS Code", "")
            if sys == "GTIN": return row.get("GTIN Code", "")
            if sys == "GMDN": return row.get("GMDN Code", "")
            return ""

        def get_description(row):
            sys = str(row.get("Code System", "")).upper()
            if sys == "SBS": return row.get("Short Description", "")
            if sys == "GTIN": return row.get("GTIN Ingredients", "")
            if sys == "GMDN": return row.get("GMDN Name", "")
            return ""

        def get_other_value(row):
            sys = str(row.get("Code System", "")).upper()
            if sys == "SBS": return row.get("SBS Code Hyphenated", "")
            if sys == "GTIN": return row.get("GTIN Strength", "")
            if sys == "GMDN": return row.get("GMDN Definition", "")
            return ""

        combined_df["service type"] = combined_df.apply(get_service_type, axis=1)
        combined_df["NPHIES Code"] = combined_df.apply(get_nphies_code, axis=1)
        combined_df["Description"] = combined_df.apply(get_description, axis=1)
        combined_df["Other Code Value"] = combined_df.apply(get_other_value, axis=1)
        combined_df["Confidence"] = combined_df.get("Confidence", "")
        combined_df["Reason"] = combined_df.get("Reasoning", "")
        
        # Define output columns
        output_cols = [
            "Service Code", 
            "Service Description", 
            "service type", 
            "NPHIES Code", 
            "Description", 
            "Other Code Value", 
            "Confidence", 
            "Reason"
        ]
        
        # Ensure columns exist
        for col in output_cols:
            if col not in combined_df.columns:
                combined_df[col] = ""
                
        # Filter to only the requested columns
        combined_df = combined_df[output_cols]
        
        # Styling
        yellow_fill = PatternFill(start_color="#FACC15", end_color="#FACC15", fill_type="solid")
        red_fill = PatternFill(start_color="#F87171", end_color="#F87171", fill_type="solid")
        
        # Write to single Excel file
        with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
            combined_df.to_excel(writer, sheet_name="Matched Results", index=False)
            
            # Apply styling
            ws = writer.sheets["Matched Results"]
            
            # Columns to highlight (1-based index)
            # 4: NPHIES Code
            # 5: Description
            # 6: Other Code Value
            highlight_cols = [4, 5, 6]
            
            for row in range(2, len(combined_df) + 2):
                # Check 'service type' value (3rd column)
                cell_value = ws.cell(row=row, column=3).value
                
                if str(cell_value).lower() == "undefined":
                    # Color entire row Red
                    for col_idx in range(1, len(output_cols) + 1):
                        ws.cell(row=row, column=col_idx).fill = red_fill
                else:
                    # Color specific columns Yellow
                    for col_idx in highlight_cols:
                        ws.cell(row=row, column=col_idx).fill = yellow_fill

        print(f"\nResults saved to {output_excel_path}")

        # Print summary statistics
        print("\n" + "=" * 60)
        print("MATCHING SUMMARY")
        print("=" * 60)
        for sheet_name, s in stats.items():
            print(f"\n{sheet_name}:")
            print(f"  Total: {s['total']}")
            print(f"  HIGH:   {s['HIGH']} ({100*s['HIGH']/s['total']:.1f}%)")
            print(f"  MEDIUM: {s['MEDIUM']} ({100*s['MEDIUM']/s['total']:.1f}%)")
            print(f"  LOW:    {s['LOW']} ({100*s['LOW']/s['total']:.1f}%)")
            print(f"  NONE:   {s['NONE']} ({100*s['NONE']/s['total']:.1f}%)")

        return all_results

    async def match_batch_async(self, input_excel_path: str, output_excel_path: str) -> list:
        """
        Process an entire mapping sheet using async for faster processing.

        Args:
            input_excel_path: Path to the input Excel file.
            output_excel_path: Path where the output Excel will be saved.

        Returns:
            List of tuples (sheet_name, DataFrame) with results.
        """
        import os
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)

        xls = pd.ExcelFile(input_excel_path)
        all_results = []
        stats = {}

        # Semaphore to limit concurrent requests
        conc = getattr(config, "ASYNC_CONCURRENCY", 1)
        semaphore = asyncio.Semaphore(conc)

        async def match_single_async(desc: str) -> dict:
            """Async wrapper for match_single with rate limiting."""
            async with semaphore:
                # Run the synchronous match in a thread pool
                loop = asyncio.get_event_loop()
                result = await loop.run_in_executor(None, self.match_single, desc)
                # Delay to avoid overwhelming the API
                delay = getattr(config, "MATCH_BATCH_DELAY", 5.0)
                await asyncio.sleep(delay)
                return result

        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df.columns = df.columns.str.strip()

            if "Service Code" not in df.columns or "Service Description" not in df.columns:
                print(f"Warning: Sheet '{sheet_name}' missing required columns, skipping")
                continue

            print(f"\nProcessing sheet: {sheet_name} ({len(df)} rows)")

            # Get all descriptions
            descriptions = df["Service Description"].tolist()

            # Process all descriptions concurrently with progress tracking
            import time
            start_time = time.time()
            completed = 0
            total = len(descriptions)
            
            async def track_progress(task, idx, desc):
                nonlocal completed
                result = await task
                completed += 1
                
                progress = (completed / total) * 100
                elapsed = time.time() - start_time
                
                desc_preview = str(desc)[:50] if desc else "(empty)"
                
                # Print progress for EVERY row to show logs in server
                if completed > 0:
                    avg_time = elapsed / completed
                    remaining = total - completed
                    eta_seconds = avg_time * remaining
                    eta_minutes = eta_seconds / 60
                    print(f"  [{sheet_name}] [{completed}/{total}] ({progress:.1f}%) | ETA: {eta_minutes:.1f}m | {desc_preview}...", flush=True)
                
                return result
            
            tasks = [track_progress(match_single_async(desc), idx, desc) for idx, desc in enumerate(descriptions)]
            results = await asyncio.gather(*tasks)

            # Extract results into lists
            matched_codes = [r.get("matched_code", "") for r in results]
            code_systems = [r.get("code_system", "") for r in results]
            matched_descriptions = [r.get("matched_description", "") for r in results]
            confidences = [r.get("confidence", "NONE") for r in results]
            reasonings = [r.get("reasoning", "") for r in results]
            
            # Extract SBS-specific fields
            sbs_code_numeric_list = [r.get("sbs_code_numeric", "") for r in results]
            sbs_short_desc_list = [r.get("short_description", "") for r in results]
            sbs_code_hyphenated_list = [r.get("sbs_code_hyphenated", "") for r in results]
            
            # Extract GTIN-specific fields
            gtin_code_list = [r.get("gtin_code", "") for r in results]
            gtin_ingredients_list = [r.get("gtin_ingredients", "") for r in results]
            gtin_strength_list = [r.get("gtin_strength", "") for r in results]
            
            # Extract GMDN-specific fields
            gmdn_code_list = [r.get("gmdn_code", "") for r in results]
            gmdn_name_list = [r.get("gmdn_name", "") for r in results]
            gmdn_definition_list = [r.get("gmdn_definition", "") for r in results]

            # Calculate stats
            sheet_stats = {"total": len(df), "HIGH": 0, "MEDIUM": 0, "LOW": 0, "NONE": 0}
            for conf in confidences:
                if conf in sheet_stats:
                    sheet_stats[conf] += 1

            # Add new columns to DataFrame in the requested order
            df["SBS Code"] = sbs_code_numeric_list
            df["Short Description"] = sbs_short_desc_list
            df["SBS Code Hyphenated"] = sbs_code_hyphenated_list
            
            df["GTIN Code"] = gtin_code_list
            df["GTIN Ingredients"] = gtin_ingredients_list
            df["GTIN Strength"] = gtin_strength_list
            
            df["GMDN Code"] = gmdn_code_list
            df["GMDN Name"] = gmdn_name_list
            df["GMDN Definition"] = gmdn_definition_list
            
            df["Matched Code"] = matched_codes
            df["Code System"] = code_systems
            df["Matched Description"] = matched_descriptions
            df["Confidence"] = confidences
            df["Reasoning"] = reasonings

            # Drop empty NPHIES columns
            cols_to_drop = ["NPHIES Code", "Description", "Other Code Value"]
            for col in cols_to_drop:
                if col in df.columns:
                    df = df.drop(columns=[col])

            all_results.append((sheet_name, df))
            stats[sheet_name] = sheet_stats

            print(f"  Completed {len(df)} rows")

        # Combine all results
        from openpyxl.styles import PatternFill

        # Combine all dataframes
        all_dfs = [df for _, df in all_results]
        if not all_dfs:
            print("No data to save.")
            return all_results
            
        combined_df = pd.concat(all_dfs, ignore_index=True)
        
        # --- Map columns to standardized output structure ---
        def get_service_type(row):
            sys = str(row.get("Code System", "")).strip()
            conf = str(row.get("Confidence", "")).upper()
            code = str(row.get("Matched Code", "")).strip()
            
            if not sys or conf == "NONE" or not code or sys.lower() == "nan":
                return "undefined"
            return sys

        def get_nphies_code(row):
            sys = str(row.get("Code System", "")).upper()
            if sys == "SBS": return row.get("SBS Code", "")
            if sys == "GTIN": return row.get("GTIN Code", "")
            if sys == "GMDN": return row.get("GMDN Code", "")
            return ""

        def get_description(row):
            sys = str(row.get("Code System", "")).upper()
            if sys == "SBS": return row.get("Short Description", "")
            if sys == "GTIN": return row.get("GTIN Ingredients", "")
            if sys == "GMDN": return row.get("GMDN Name", "")
            return ""

        def get_other_value(row):
            sys = str(row.get("Code System", "")).upper()
            if sys == "SBS": return row.get("SBS Code Hyphenated", "")
            if sys == "GTIN": return row.get("GTIN Strength", "")
            if sys == "GMDN": return row.get("GMDN Definition", "")
            return ""

        combined_df["service type"] = combined_df.apply(get_service_type, axis=1)
        combined_df["NPHIES Code"] = combined_df.apply(get_nphies_code, axis=1)
        combined_df["Description"] = combined_df.apply(get_description, axis=1)
        combined_df["Other Code Value"] = combined_df.apply(get_other_value, axis=1)
        combined_df["Confidence"] = combined_df.get("Confidence", "")
        combined_df["Reason"] = combined_df.get("Reasoning", "")
        
        # Define output columns
        output_cols = [
            "Service Code", 
            "Service Description", 
            "service type", 
            "NPHIES Code", 
            "Description", 
            "Other Code Value", 
            "Confidence", 
            "Reason"
        ]
        
        # Ensure columns exist
        for col in output_cols:
            if col not in combined_df.columns:
                combined_df[col] = ""
                
        # Filter to only the requested columns
        combined_df = combined_df[output_cols]
        
        # Styling
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Write to single Excel file
        with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
            combined_df.to_excel(writer, sheet_name="Matched Results", index=False)
            
            # Apply styling
            ws = writer.sheets["Matched Results"]
            
            # Columns to highlight (1-based index)
            # 4: NPHIES Code
            # 5: Description
            # 6: Other Code Value
            highlight_cols = [4, 5, 6]
            
            for row in range(2, len(combined_df) + 2):
                # Check 'service type' value (3rd column)
                cell_value = ws.cell(row=row, column=3).value
                
                if str(cell_value).lower() == "undefined":
                    # Color entire row Red
                    for col_idx in range(1, len(output_cols) + 1):
                        ws.cell(row=row, column=col_idx).fill = red_fill
                else:
                    # Color specific columns Yellow
                    for col_idx in highlight_cols:
                        ws.cell(row=row, column=col_idx).fill = yellow_fill

        print(f"\nResults saved to {output_excel_path}")

        # Print summary
        print("\n" + "=" * 60)
        print("MATCHING SUMMARY")
        print("=" * 60)
        for sheet_name, s in stats.items():
            print(f"\n{sheet_name}:")
            print(f"  Total: {s['total']}")
            print(f"  HIGH:   {s['HIGH']} ({100*s['HIGH']/s['total']:.1f}%)")
            print(f"  MEDIUM: {s['MEDIUM']} ({100*s['MEDIUM']/s['total']:.1f}%)")
            print(f"  LOW:    {s['LOW']} ({100*s['LOW']/s['total']:.1f}%)")
            print(f"  NONE:   {s['NONE']} ({100*s['NONE']/s['total']:.1f}%)")

        return all_results


if __name__ == "__main__":
    # Quick test
    print("Initializing CodeMatcher...")
    matcher = CodeMatcher()

    print("\nTesting single match: 'Cisternal puncture'")
    result = matcher.match_single("Cisternal puncture")
    print(json.dumps(result, indent=2))

    print("\nTesting single match: 'BEPANTHEN'")
    result = matcher.match_single("BEPANTHEN")
    print(json.dumps(result, indent=2))

    print("\nTesting single match: 'vitamin D3 IVD calibrator'")
    result = matcher.match_single("vitamin D3 IVD calibrator")
    print(json.dumps(result, indent=2))
